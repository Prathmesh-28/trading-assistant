"""screener.in adapter — isolated data source (mirrors the groww_adapter
import boundary). OFF by default (SCREENER_SCRAPE_ENABLED); the operator of
this personal tool owns the terms-of-use decision for their own account.

Uses screener.in's own **Excel export** (the robust, structured path), not
HTML scraping:
  1. GET  /api/company/search/?q=SYMBOL      -> resolve to the company URL
  2. GET  the company page                   -> warehouse id from the export
                                                form action
  3. POST /user/company/export/<id>/         -> the .xlsx report
  4. parse the "Data Sheet" tab (a stable, versioned layout) into a clean dict

Why export over HTML: the Data Sheet is a fixed schema ("LATEST VERSION 2.1"),
so it doesn't drift when the website's HTML changes. The export requires a
logged-in session, so the operator supplies their OWN cookie (SCREENER_COOKIE);
without it the POST 403s and callers fall back to yfinance.

Built defensively: process-wide rate limit + jitter, long disk cache, 403/429
cooldown, label-scan parser that fails soft to None, provenance stamp.
"""

from __future__ import annotations

import io
import json
import logging
import os
import random
import re
import time
from pathlib import Path

log = logging.getLogger("screener")

_BASE = "https://www.screener.in"
_CACHE_DIR = Path(os.getenv("SCREENER_CACHE_DIR", "/tmp/screener_cache"))
_TTL = float(os.getenv("SCREENER_CACHE_TTL_H", "18")) * 3600
_MIN_INTERVAL = float(os.getenv("SCREENER_MIN_INTERVAL", "4"))
_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

_last_request = [0.0]
_blocked_until = [0.0]


def enabled() -> bool:
    on = os.getenv("SCREENER_SCRAPE_ENABLED", "").strip().lower() in ("1", "true", "yes", "on")
    if on and not os.getenv("SCREENER_COOKIE", "").strip():
        log.warning("SCREENER_SCRAPE_ENABLED but no SCREENER_COOKIE — the Excel "
                    "export needs a logged-in session; will fall back to yfinance")
    return on


def _cache_path(symbol: str) -> Path:
    return _CACHE_DIR / f"{symbol.upper()}.json"


def _read_cache(symbol: str) -> dict | None:
    p = _cache_path(symbol)
    try:
        if p.exists() and time.time() - p.stat().st_mtime < _TTL:
            return json.loads(p.read_text())
    except Exception:  # noqa: BLE001
        pass
    return None


def _write_cache(symbol: str, data: dict) -> None:
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        _cache_path(symbol).write_text(json.dumps(data))
    except Exception:  # noqa: BLE001
        pass


def _throttle() -> None:
    wait = _MIN_INTERVAL - (time.time() - _last_request[0])
    if wait > 0:
        time.sleep(wait + random.uniform(0.2, 0.8))
    _last_request[0] = time.time()


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _cagr(first, last_, n):
    if not (first and last_) or first <= 0 or last_ <= 0 or n <= 0:
        return None
    return round(((last_ / first) ** (1 / n) - 1) * 100, 1)


def parse_datasheet(xlsx_bytes: bytes) -> dict | None:
    """Parse the ENTIRE screener 'Data Sheet' tab (schema v2.x) — every P&L,
    balance-sheet and cash-flow line across all years — then compute the full
    screener-style ratio suite (ROCE, OPM/NPM, working-capital days, Piotroski
    F-Score, Altman Z, CAGRs, FCF). Label-scan so layout shifts still resolve;
    returns None on real drift so callers fall back to yfinance."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        log.warning("screener export not a valid xlsx: %s", e)
        return None
    if "Data Sheet" not in wb.sheetnames:
        log.warning("screener drift: no 'Data Sheet' tab — failing soft")
        return None
    rows = [[c.value for c in row] for row in wb["Data Sheet"].iter_rows()]

    def label_row(label, start=0, end=None):
        end = end if end is not None else len(rows)
        lab = label.lower()
        for i in range(start, min(end, len(rows))):
            a = rows[i][0]
            if isinstance(a, str) and a.strip().lower() == lab:
                return i
        return None

    def cell(label, start=0, end=None):
        i = label_row(label, start, end)
        return _f(rows[i][1]) if i is not None and len(rows[i]) > 1 else None

    def series(label, start, end=None):
        i = label_row(label, start, end)
        return [_f(v) for v in rows[i][1:] if _f(v) is not None] if i is not None else []

    def last(xs):
        return xs[-1] if xs else None

    def pct(n, d):
        return round(n / d * 100, 1) if n is not None and d else None

    pl = label_row("PROFIT & LOSS") or 0
    qs = label_row("Quarters") or (pl + 25)
    bs = label_row("BALANCE SHEET") or (qs + 15)
    cf = label_row("CASH FLOW:") or (bs + 25)

    def raw_series(label, start, end):
        i = label_row(label, start, end)
        return rows[i][1:] if i is not None else []

    # --- full annual P&L ---
    report_years = []
    for v in raw_series("Report Date", pl, qs):
        m = re.match(r"(\d{4})", str(v)) if v is not None else None
        report_years.append(m.group(1) if m else None)
    sales = series("Sales", pl, qs)
    net_profit = series("Net profit", pl, qs)
    op_profit = series("Operating Profit", pl, qs)
    other_income = series("Other Income", pl, qs)
    depreciation = series("Depreciation", pl, qs)
    interest = series("Interest", pl, qs)
    pbt = series("Profit before tax", pl, qs)
    tax = series("Tax", pl, qs)
    dividend = series("Dividend Amount", pl, qs)
    eps_row = series("EPS", pl, qs)
    if not sales or not net_profit:
        log.warning("screener drift: Sales/Net-profit empty — failing soft")
        return None

    # --- full balance sheet ---
    equity = series("Equity Share Capital", bs, cf)
    reserves = series("Reserves", bs, cf)
    borrowings = series("Borrowings", bs, cf)
    other_liab = series("Other Liabilities", bs, cf)
    net_block = series("Net Block", bs, cf)
    investments = series("Investments", bs, cf)
    other_assets = series("Other Assets", bs, cf)
    receivables = series("Receivables", bs, cf) or series("Debtors", bs, cf)
    inventory = series("Inventory", bs, cf)
    cash_bank = series("Cash & Bank", bs, cf)
    shares = series("No. of Equity Shares", bs, cf)

    # --- cash flow ---
    cfo = series("Cash from Operating Activity", cf)
    cfi = series("Cash from Investing Activity", cf)
    cff = series("Cash from Financing Activity", cf)

    name = rows[0][1] if rows and len(rows[0]) > 1 and isinstance(rows[0][1], str) else None
    price = cell("Current Price", 0, pl)
    mcap = cell("Market Capitalization", 0, pl)
    face_value = cell("Face Value", 0, pl)

    latest_np = last(net_profit)
    eq_total = [(equity[i] if i < len(equity) else 0) + (reserves[i] if i < len(reserves) else 0)
                for i in range(max(len(equity), len(reserves)))]
    latest_eq = last(eq_total)
    latest_borrow = last(borrowings)

    # annual Operating Profit isn't a direct row — reconstruct it:
    # OP = PBT + Interest + Depreciation - Other Income
    def _at(xs, i):
        return xs[i] if i < len(xs) else 0
    op_annual = [pbt[i] + _at(interest, i) + _at(depreciation, i) - _at(other_income, i)
                 for i in range(len(pbt))] if pbt else []
    op_profit = op_profit or op_annual

    # ROCE = EBIT / (equity + borrowings) ; EBIT ≈ PBT + interest
    ebit = (last(pbt) or 0) + (last(interest) or 0) if pbt else None
    capital_employed = (latest_eq or 0) + (latest_borrow or 0)
    roce = pct(ebit, capital_employed) if ebit and capital_employed else None

    # FCF ≈ CFO − capex(≈ change in net block, approx via investing) ; use CFO as proxy floor
    fcf = last(cfo)

    # working-capital days (latest)
    debtor_days = (round(last(receivables) / last(sales) * 365, 0)
                   if last(receivables) and last(sales) else None)
    inventory_days = (round(last(inventory) / last(sales) * 365, 0)
                      if last(inventory) and last(sales) else None)

    # Piotroski F-Score (9 pts) — needs 2 consecutive years
    piotroski = None
    if len(net_profit) >= 2 and len(eq_total) >= 2:
        f = 0
        f += 1 if latest_np > 0 else 0                                    # 1 positive NI
        f += 1 if fcf and fcf > 0 else 0                                  # 2 positive CFO
        f += 1 if len(net_profit) >= 2 and pct(latest_np, latest_eq) and \
            pct(net_profit[-2], eq_total[-2] or 1) and \
            pct(latest_np, latest_eq) > pct(net_profit[-2], eq_total[-2] or 1) else 0  # 3 rising ROA
        f += 1 if fcf and fcf > latest_np else 0                          # 4 CFO > NI (accruals)
        f += 1 if len(borrowings) >= 2 and (last(borrowings) or 0) <= (borrowings[-2] or 0) else 0  # 5 lower leverage
        f += 1 if len(op_profit) >= 2 and len(sales) >= 2 and \
            pct(last(op_profit), last(sales)) and pct(op_profit[-2], sales[-2] or 1) and \
            pct(last(op_profit), last(sales)) > pct(op_profit[-2], sales[-2] or 1) else 0  # 6 rising margin
        f += 1 if len(sales) >= 2 and last(sales) > (sales[-2] or 0) else 0  # 7 asset turnover proxy (sales growth)
        f += 1 if len(shares) < 2 or (last(shares) or 0) <= (shares[-2] or 0) * 1.02 else 0  # 8 no big dilution
        f += 1 if latest_np > (net_profit[-2] or 0) else 0                # 9 profit growth
        piotroski = f

    # Altman Z (approx, manufacturing form) — working capital, retained (reserves),
    # EBIT, market cap, sales vs total assets
    total_assets = last(other_assets) is not None and (
        (last(net_block) or 0) + (last(investments) or 0) + (last(other_assets) or 0)) or None
    altman_z = None
    if total_assets and total_assets > 0:
        wc = (last(cash_bank) or 0) + (last(receivables) or 0) + (last(inventory) or 0) - (last(other_liab) or 0)
        z = (1.2 * wc / total_assets
             + 1.4 * (last(reserves) or 0) / total_assets
             + 3.3 * (ebit or 0) / total_assets
             + 0.6 * (mcap or 0) / ((last(borrowings) or 0) + (last(other_liab) or 1))
             + 1.0 * (last(sales) or 0) / total_assets)
        altman_z = round(z, 2)

    out = {
        "source": "screener.in",
        "name": name,
        "face_value": face_value,
        "market_cap": mcap,
        "pe": round(mcap / latest_np, 2) if mcap and latest_np else None,
        "roe_pct": pct(latest_np, latest_eq),
        "roce_pct": roce,
        "debt_to_equity": round(latest_borrow / latest_eq, 2) if latest_borrow is not None and latest_eq else None,
        "opm_pct": pct(last(op_profit), last(sales)),
        "npm_pct": pct(latest_np, last(sales)),
        "interest_coverage": round(ebit / last(interest), 1) if ebit and last(interest) else None,
        "payout_ratio_pct": pct(last(dividend), latest_np),
        "dividend_yield_pct": None,
        "fcf": fcf,
        "fcf_yield_pct": pct(fcf, mcap),
        "debtor_days": debtor_days,
        "inventory_days": inventory_days,
        "revenue_cagr_pct": _cagr(sales[0], sales[-1], len(sales) - 1),
        "pat_cagr_pct": _cagr(net_profit[0], net_profit[-1], len(net_profit) - 1),
        "sales_cagr_5y": _cagr(sales[-6], sales[-1], 5) if len(sales) >= 6 else None,
        "pat_cagr_5y": _cagr(net_profit[-6], net_profit[-1], 5) if len(net_profit) >= 6 else None,
        "piotroski_score": piotroski,
        "altman_z": altman_z,
        "eps": round(eps_row[-1], 2) if eps_row else None,
        "cfo_latest": last(cfo),
        "years": [],
        "cashflow": {"operating": cfo[-5:], "investing": cfi[-5:], "financing": cff[-5:]},
    }
    n = min(len(sales), len(net_profit))
    prev = None
    for i in range(n):
        g = round((sales[i] / prev - 1) * 100, 1) if prev else None
        opm = pct(op_profit[i], sales[i]) if i < len(op_profit) else None
        yr = report_years[i] if i < len(report_years) else None
        out["years"].append({"year": yr, "revenue": sales[i], "net_income": net_profit[i],
                             "revenue_growth_pct": g, "opm_pct": opm})
        prev = sales[i]
    return out


def fetch(symbol: str) -> dict | None:
    if not enabled():
        return None
    cached = _read_cache(symbol)
    if cached is not None:
        return cached
    if time.time() < _blocked_until[0]:
        return None
    cookie = os.getenv("SCREENER_COOKIE", "").strip()
    if not cookie:
        return None  # export needs a session; fall back cleanly
    try:
        import httpx

        with httpx.Client(timeout=20, follow_redirects=True,
                          headers={"User-Agent": _UA, "Cookie": cookie}) as client:
            _throttle()
            search = client.get(f"{_BASE}/api/company/search/", params={"q": symbol})
            if search.status_code in (403, 429):
                _blocked_until[0] = time.time() + 1800
                log.warning("screener search %s -> %s; 30min cooldown", symbol, search.status_code)
                return None
            hits = search.json() if search.status_code == 200 else []
            url = next((h["url"] for h in hits
                        if h.get("url", "").upper().rstrip("/").endswith(symbol.upper())), None)
            url = url or (hits[0]["url"] if hits else None)
            if not url:
                log.warning("screener: symbol %s not found", symbol)
                return None

            _throttle()
            page = client.get(f"{_BASE}{url}")
            m = re.search(r"/user/company/export/(\d+)/", page.text)
            csrf = client.cookies.get("csrftoken", "")
            if not m:
                log.warning("screener drift: export id not found on %s page", symbol)
                return None
            wid = m.group(1)

            _throttle()
            export = client.post(
                f"{_BASE}/user/company/export/{wid}/",
                data={"csrfmiddlewaretoken": csrf, "next": url},
                headers={"Referer": f"{_BASE}{url}", "Content-Type": "application/x-www-form-urlencoded"},
            )
            if export.status_code in (403, 429):
                _blocked_until[0] = time.time() + 1800
                log.warning("screener export %s -> %s; 30min cooldown", symbol, export.status_code)
                return None
            if export.status_code != 200:
                log.warning("screener export %s -> HTTP %s", symbol, export.status_code)
                return None
            data = parse_datasheet(export.content)
            if data:
                _write_cache(symbol, data)
            return data
    except Exception as e:  # noqa: BLE001 — never break the fundamentals card
        log.warning("screener fetch %s failed: %s", symbol, e)
        return None


# ---------------------------------------------------------------- page extras
# The PUBLIC company page carries data the Excel export doesn't: shareholding
# pattern, quarterly results grid, screener's official CAGR tables, machine-
# generated pros/cons, working-capital days, peers, documents/concalls.
# No login needed — this path works without SCREENER_COOKIE.

def _table_rows(section) -> "tuple[dict, list]":
    """{row-label-lower: [floats]} + column headers, from a section's
    data table. Percent/comma noise stripped; blanks become None."""
    tbl = section.select_one("table.data-table") or section.select_one("table")
    if tbl is None:
        return {}, []
    headers = [th.get_text(strip=True) for th in tbl.select("thead th")][1:]
    rows = {}
    for tr in tbl.select("tbody tr") or tbl.select("tr"):
        cells = tr.select("td")
        if len(cells) < 2:
            continue
        label = cells[0].get_text(" ", strip=True)
        label = re.sub(r"\s*\+\s*$", "", label).strip().lower()
        vals = []
        for c in cells[1:]:
            t = c.get_text(strip=True).replace(",", "").replace("%", "")
            vals.append(_f(t))
        if label:
            rows[label] = vals
    return rows, headers


def parse_page(html: str) -> dict | None:
    """Everything screener.in shows on a company page, as one dict. Label-scan
    everywhere; a missing section just yields Nones (fail-soft)."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    top = soup.select_one("#top-ratios")
    if top is None:
        log.warning("screener page drift: #top-ratios missing — failing soft")
        return None

    ratios = {}
    for li in top.select("li"):
        nm = li.select_one(".name")
        if not nm:
            continue
        val = li.select_one(".nowrap.value") or li.select_one(".value") or li
        txt = val.get_text(" ", strip=True)
        ratios[nm.get_text(strip=True).lower()] = txt

    def rnum(*keys, pick=0):
        for k in keys:
            for name, txt in ratios.items():
                if k in name:
                    nums = re.findall(r"-?[\d,]+\.?\d*", txt)
                    if len(nums) > pick:
                        return _f(nums[pick].replace(",", ""))
        return None

    def sect(sid):
        return soup.select_one(f"section#{sid}") or soup.select_one(f"#{sid}")

    out = {
        "source": "screener.in",
        "current_price": rnum("current price"),
        "market_cap": rnum("market cap"),
        "pe": rnum("stock p/e", "p/e"),
        "book_value": rnum("book value"),
        "dividend_yield_pct": rnum("dividend yield"),
        "roce_pct": rnum("roce"),
        "roe_pct": rnum("roe"),
        "face_value": rnum("face value"),
        "high_52w": rnum("high / low"),
        "low_52w": rnum("high / low", pick=1),
    }
    out["pb"] = (round(out["current_price"] / out["book_value"], 2)
                 if out["current_price"] and out["book_value"] else None)

    out["pros"] = [li.get_text(" ", strip=True) for li in soup.select(".pros li")][:8]
    out["cons"] = [li.get_text(" ", strip=True) for li in soup.select(".cons li")][:8]

    # quarterly results grid (last 13 quarters)
    q = sect("quarters")
    if q:
        rows, headers = _table_rows(q)
        out["quarters"] = {
            "headers": headers,
            "sales": rows.get("sales", []),
            "operating_profit": rows.get("operating profit", []),
            "opm_pct": rows.get("opm %", []),
            "net_profit": rows.get("net profit", []),
            "eps": rows.get("eps in rs", []),
        }

    # official CAGR mini-tables (Compounded Sales/Profit Growth, Stock Price
    # CAGR, Return on Equity) inside the P&L section
    growth = {}
    pl = sect("profit-loss")
    if pl:
        for t in pl.select("table.ranges-table"):
            th = t.select_one("th")
            title = th.get_text(" ", strip=True).rstrip(":").lower() if th else ""
            entries = {}
            for tr in t.select("tr"):
                tds = tr.select("td")
                if len(tds) == 2:
                    k = tds[0].get_text(strip=True).rstrip(":").lower()
                    entries[k] = _f(tds[1].get_text(strip=True).replace("%", ""))
            if "sales" in title:
                growth["sales_cagr"] = entries
            elif "profit" in title:
                growth["profit_cagr"] = entries
            elif "price" in title:
                growth["price_cagr"] = entries
            elif "equity" in title:
                growth["roe_trend"] = entries
    out["growth"] = growth

    # ratios trend (Debtor Days / Working Capital Days / ROCE %)
    rt = sect("ratios")
    if rt:
        rows, headers = _table_rows(rt)
        out["ratio_trends"] = {
            "headers": headers,
            "debtor_days": rows.get("debtor days", []),
            "working_capital_days": rows.get("working capital days", []),
            "roce_pct": rows.get("roce %", []),
        }
        wcd = rows.get("working capital days", [])
        out["working_capital_days"] = wcd[-1] if wcd else None

    # shareholding pattern (quarterly % + shareholder count)
    sh = sect("shareholding")
    if sh:
        rows, headers = _table_rows(sh)
        out["shareholding"] = {
            "headers": headers,
            "promoters": rows.get("promoters", []),
            "fiis": rows.get("fiis", []),
            "diis": rows.get("diis", []),
            "public": rows.get("public", []),
            "shareholders": rows.get("no. of shareholders", []),
        }
        for key, row in (("promoter_pct", "promoters"), ("fii_pct", "fiis"),
                         ("dii_pct", "diis"), ("public_pct", "public")):
            vals = rows.get(row, [])
            out[key] = vals[-1] if vals else None
        proms = rows.get("promoters", [])
        out["promoter_change_pct"] = (round(proms[-1] - proms[0], 2)
                                      if len(proms) >= 2 and proms[-1] is not None
                                      and proms[0] is not None else None)

    # documents: annual reports + concall links (URLs only, never re-hosted)
    docs = {"annual_reports": [], "concalls": []}
    dsec = sect("documents")
    if dsec:
        ar_box = dsec.select_one(".annual-reports")
        for a in (ar_box.select("a") if ar_box else [])[:12]:
            href = a.get("href", "")
            label = " ".join(a.get_text(" ", strip=True).split())[:40]
            if href.startswith("http"):
                docs["annual_reports"].append({"label": label, "url": href})
        for li in dsec.select(".concalls li, ul.show-more-box li")[:8]:
            date = li.find(string=lambda t: t and t.strip())
            links = {a.get_text(strip=True).lower(): a.get("href", "")
                     for a in li.select("a") if a.get("href", "").startswith("http")}
            if links:
                docs["concalls"].append({"date": str(date).strip()[:12], **links})
    out["documents"] = docs

    # company id for the peers API
    m = re.search(r'data-company-id="(\d+)"', html)
    out["company_id"] = m.group(1) if m else None
    return out


def _parse_peers(html: str) -> list:
    """Peer-comparison table parser — RESERVED, not called yet: screener's
    peers API needs a logged-in session (public probe 404s). Wire it up when
    a SCREENER_COOKIE flow lands, or build peers from our own universe."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    tbl = soup.select_one("table")
    if tbl is None:
        return []
    headers = [th.get_text(" ", strip=True).lower() for th in tbl.select("thead th")]

    def col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    ci = {"name": col("name"), "cmp": col("cmp"), "pe": col("p/e"),
          "mcap": col("mar cap"), "roce": col("roce"), "np_qtr": col("np qtr"),
          "sales_qtr": col("sales qtr")}
    peers = []
    for tr in tbl.select("tbody tr"):
        tds = tr.select("td")
        if len(tds) < 3 or "median" in tds[0].get_text(strip=True).lower():
            continue
        row = {}
        for key, idx in ci.items():
            if idx is None or idx >= len(tds):
                continue
            txt = tds[idx].get_text(" ", strip=True)
            row[key] = txt if key == "name" else _f(txt.replace(",", ""))
        if row.get("name"):
            peers.append(row)
    return peers[:10]


def fetch_page(symbol: str) -> dict | None:
    """Public company-page extras (shareholding, quarters, CAGRs,
    pros/cons, documents). Cached like the export; no cookie required."""
    if not enabled():
        return None
    cached = _read_cache(f"{symbol}_page")
    if cached is not None:
        return cached
    if time.time() < _blocked_until[0]:
        return None
    try:
        import httpx

        headers = {"User-Agent": _UA}
        cookie = os.getenv("SCREENER_COOKIE", "").strip()
        if cookie:
            headers["Cookie"] = cookie
        with httpx.Client(timeout=20, follow_redirects=True, headers=headers) as client:
            _throttle()
            resp = client.get(f"{_BASE}/company/{symbol.upper()}/")
            if resp.status_code in (403, 429):
                _blocked_until[0] = time.time() + 1800
                log.warning("screener page %s -> %s; 30min cooldown", symbol, resp.status_code)
                return None
            if resp.status_code != 200:
                log.warning("screener page %s -> HTTP %s", symbol, resp.status_code)
                return None
            data = parse_page(resp.text)
            if not data:
                return None
            _write_cache(f"{symbol}_page", data)
            return data
    except Exception as e:  # noqa: BLE001
        log.warning("screener page fetch %s failed: %s", symbol, e)
        return None
